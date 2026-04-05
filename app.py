# -*- coding: utf-8 -*-
import os
import logging
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# =========================================
# CONFIGURACION DE LOGGING
# =========================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('app_audit.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# =========================================
# CONFIGURACION GENERAL
# =========================================

APP_NAME = "SEPTEM"

# Limites de seguridad
MAX_CARACTERES = 10000
MAX_TAMAÑO_ARCHIVO = 5 * 1024 * 1024  # 5 MB
TAMAÑO_MAXIMO_NOMBRE = 255
EXTENSIONES_PERMITIDAS = {'.xlsx', '.txt'}

ALFABETO = (
    "abcdefghijklmnopqrstuvwxyz"
    "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    "0123456789"
    " áéíóúÁÉÍÓÚñÑ"
    ".,;:-_¿?¡!\"'()[]{}<>/@#%&+=*\\|~^`$"
    "\n\t"
)

COLUMNAS_GRILLA = 12
CLAVE_AUREA = "1.618"
PHI = 1.618
CLAVE_BASE = int(PHI * 1000)  # 1618


# =========================================
# FUNCIONES DE VALIDACION
# =========================================

def validar_texto(texto):
    """Valida que el texto cumpla con los limites de seguridad."""
    if not isinstance(texto, str):
        logger.warning(f"Intento de validar texto no string: {type(texto)}")
        return False, "El texto debe ser una cadena."
    
    if len(texto) > MAX_CARACTERES:
        logger.warning(f"Texto excede limite maximo: {len(texto)} > {MAX_CARACTERES}")
        return False, f"El texto no puede exceder {MAX_CARACTERES} caracteres."
    
    return True, None


def validar_nombre_archivo(nombre):
    """Valida que el nombre del archivo sea seguro."""
    if not isinstance(nombre, str) or not nombre.strip():
        logger.warning("Nombre de archivo inválido o vacío")
        return False, "El nombre no puede estar vacío."
    
    if len(nombre) > TAMAÑO_MAXIMO_NOMBRE:
        logger.warning(f"Nombre de archivo excede limite: {len(nombre)}")
        return False, f"El nombre no puede exceder {TAMAÑO_MAXIMO_NOMBRE} caracteres."
    
    # Caracteres prohibidos en nombres de archivo
    caracteres_prohibidos = '<>:"|?*\\'
    if any(char in nombre for char in caracteres_prohibidos):
        logger.warning(f"Nombre contiene caracteres prohibidos: {nombre}")
        return False, "El nombre contiene caracteres no permitidos."
    
    return True, None


def validar_ruta_archivo(ruta):
    """Valida que la ruta sea segura."""
    try:
        ruta_path = Path(ruta).resolve()
        
        # Si el archivo ya existe, verificar que sea un archivo
        if ruta_path.exists() and not ruta_path.is_file():
            logger.warning(f"Ruta no es un archivo: {ruta}")
            return False, "La ruta debe ser un archivo, no un directorio."
        
        # Verificar que la carpeta padre existe o se puede crear
        carpeta_padre = ruta_path.parent
        if not carpeta_padre.exists():
            logger.warning(f"Carpeta padre no existe: {carpeta_padre}")
            # Intentar crear la carpeta
            try:
                carpeta_padre.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                logger.error(f"No se puede crear carpeta: {e}")
                return False, f"No se puede acceder a la carpeta: {str(e)}"
        
        return True, None
    except (OSError, ValueError) as e:
        logger.error(f"Error validando ruta: {e}")
        return False, f"Ruta inválida: {str(e)}"


# =========================================
# LOGICA DE CIFRADO
# =========================================

def generar_fibonacci(n):
    fib = []
    a, b = 0, 1
    for _ in range(n):
        fib.append(a)
        a, b = b, a + b
    return fib


def texto_a_numeros(texto):
    """Convierte texto a números con validación de seguridad."""
    valido, error = validar_texto(texto)
    if not valido:
        logger.warning(f"Texto rechazado en conversion: {error}")
        return []
    
    tamaño = len(ALFABETO)
    fib = generar_fibonacci(len(texto))
    numeros = []

    for i, caracter in enumerate(texto):
        if caracter in ALFABETO:
            indice_original = ALFABETO.index(caracter)
            indice_cifrado = (indice_original + fib[i] + CLAVE_BASE) % tamaño
            numeros.append(indice_cifrado)
        else:
            logger.warning(f"Caracter no soportado: {repr(caracter)}")
            numeros.append(-1)

    logger.info(f"Conversión completada: {len(texto)} caracteres -> {len(numeros)} números")
    return numeros


def numeros_a_texto(numeros, clave_ingresada):
    """Convierte números a texto con validación de clave."""
    if not isinstance(numeros, list):
        logger.error("Error: numeros no es una lista")
        return False, ""
    
    if not isinstance(clave_ingresada, str):
        logger.warning("Clave no válida: no es string")
        return False, ""
    
    tamaño = len(ALFABETO)
    fib = generar_fibonacci(len(numeros))

    clave_limpia = clave_ingresada.strip()
    
    if clave_limpia == CLAVE_AUREA:
        resultado = []
        for i, numero in enumerate(numeros):
            if numero == -1:
                resultado.append("?")
            else:
                try:
                    indice_original = (numero - fib[i] - CLAVE_BASE) % tamaño
                    resultado.append(ALFABETO[indice_original])
                except (IndexError, TypeError) as e:
                    logger.error(f"Error decifrando número {numero}: {e}")
                    resultado.append("?")
        
        texto_resultado = "".join(resultado)
        logger.info("Decodificación exitosa con clave correcta")
        return True, texto_resultado

    # Clave incorrecta - mostrar mensaje de error
    logger.warning("Intento de decodificación con clave incorrecta")
    return False, "La proporción no es divina"


def numeros_a_matriz(numeros, columnas=COLUMNAS_GRILLA):
    matriz = []
    fila = []
    for n in numeros:
        fila.append(n)
        if len(fila) == columnas:
            matriz.append(fila)
            fila = []

    if fila:
        while len(fila) < columnas:
            fila.append("")
        matriz.append(fila)

    return matriz


def matriz_a_numeros(matriz):
    """Convierte matriz a números con validación de tipos."""
    if not isinstance(matriz, list):
        logger.error("Matriz no válida: no es una lista")
        return []
    
    numeros = []
    for i, fila in enumerate(matriz):
        if not isinstance(fila, (list, tuple)):
            logger.warning(f"Fila {i} no es lista/tupla: {type(fila)}")
            continue
        
        for j, valor in enumerate(fila):
            if valor is None or valor == "":
                continue
            
            try:
                # Validar que el valor puede ser convertido a int
                num = int(valor)
                # Validación adicional: el número debe ser válido para el cifrado
                if 0 <= num < len(ALFABETO):
                    numeros.append(num)
                else:
                    logger.warning(f"Número fuera de rango en posición ({i},{j}): {num}")
            except (ValueError, TypeError) as e:
                logger.warning(f"No se puede convertir valor en ({i},{j}): {repr(valor)}")
    
    return numeros


# =========================================
# EXCEL
# =========================================

def escribir_excel(ruta, numeros):
    """Escribe números a archivo Excel con validación."""
    valido, error = validar_ruta_archivo(ruta)
    if not valido:
        logger.error(f"Ruta de escritura rechazada: {error}")
        raise ValueError(error)
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "SECUENCIA"

        matriz = numeros_a_matriz(numeros)

        for i, fila in enumerate(matriz, start=1):
            for j, valor in enumerate(fila, start=1):
                ws.cell(row=i, column=j, value=valor)

        wb.save(ruta)
        logger.info(f"Archivo Excel guardado exitosamente: {ruta}")
        
    except Exception as e:
        logger.error(f"Error escribiendo archivo Excel: {type(e).__name__}: {str(e)}")
        raise


def construir_nombre_archivo(nombre_base):
    """Construye nombre de archivo seguro con timestamp."""
    nombre_limpio = nombre_base.strip()
    marca_tiempo = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Eliminar extensiones si las hay
    nombre_sin_ext = Path(nombre_limpio).stem
    
    nombre_final = f"{nombre_sin_ext}_{marca_tiempo}.xlsx"
    logger.info(f"Nombre de archivo construido: {nombre_final}")
    return nombre_final


def guardar_como_nuevo(numeros):
    """Guarda nuevo archivo con validaciones de seguridad."""
    if not numeros:
        logger.warning("Intento de guardar contenido vacío")
        messagebox.showwarning("Aviso", "No hay contenido para guardar.")
        return None

    nombre_base = simpledialog.askstring("Guardar", "Ingrese el nombre del archivo:")
    if not nombre_base:
        logger.info("Operación de guardado cancelada por usuario")
        return None

    valido, error_msg = validar_nombre_archivo(nombre_base)
    if not valido:
        logger.warning(f"Nombre de archivo rechazado: {error_msg}")
        messagebox.showerror("Error", error_msg)
        return None

    nombre_final = construir_nombre_archivo(nombre_base)

    ruta = filedialog.asksaveasfilename(
        title="Guardar archivo",
        defaultextension=".xlsx",
        initialfile=nombre_final,
        filetypes=[("Archivos Excel", "*.xlsx")]
    )

    if not ruta:
        logger.info("Operación de guardado cancelada")
        return None

    try:
        escribir_excel(ruta, numeros)
        logger.info(f"Archivo guardado exitosamente: {ruta}")
        return ruta
    except Exception as e:
        logger.error(f"Error al guardar archivo: {e}")
        messagebox.showerror("Error", f"No se pudo guardar el archivo: {str(e)}")
        return None


def sobrescribir_archivo_existente(ruta, numeros):
    """Sobrescribe archivo existente con validaciones."""
    if not ruta:
        logger.warning("Intento de sobrescribir sin ruta")
        return None
    
    try:
        escribir_excel(ruta, numeros)
        logger.info(f"Archivo sobrescrito: {ruta}")
        return ruta
    except Exception as e:
        logger.error(f"Error sobrescribiendo archivo: {e}")
        messagebox.showerror("Error", f"No se pudo sobrescribir el archivo: {str(e)}")
        return None


def abrir_desde_excel():
    """Abre archivo Excel con validaciones de seguridad."""
    ruta = filedialog.askopenfilename(
        title="Abrir archivo",
        filetypes=[("Archivos Excel", "*.xlsx")]
    )

    if not ruta:
        logger.info("Operación de apertura cancelada")
        return None, None

    try:
        # Validación de seguridad
        valido, error = validar_ruta_archivo(ruta)
        if not valido:
            logger.error(f"Ruta rechazada: {error}")
            messagebox.showerror("Error", error)
            return None, None
        
        # Verificar tamaño del archivo
        tamaño_archivo = os.path.getsize(ruta)
        if tamaño_archivo > MAX_TAMAÑO_ARCHIVO:
            logger.warning(f"Archivo muy grande: {tamaño_archivo} bytes")
            messagebox.showerror("Error", f"El archivo es demasiado grande (máximo {MAX_TAMAÑO_ARCHIVO} bytes)")
            return None, None

        # Cargar y validar archivo Excel
        try:
            wb = load_workbook(ruta)
        except InvalidFileException as e:
            logger.error(f"Archivo no válido: {e}")
            messagebox.showerror("Error", "El archivo no es un archivo Excel válido.")
            return None, None
        except Exception as e:
            logger.error(f"Error abriendo archivo Excel: {e}")
            messagebox.showerror("Error", f"No se pudo abrir el archivo: {str(e)}")
            return None, None

        # Obtener hoja correcta
        try:
            ws = wb["SECUENCIA"] if "SECUENCIA" in wb.sheetnames else wb.active
        except Exception as e:
            logger.error(f"Error al acceder a hoja: {e}")
            messagebox.showerror("Error", "No se pudo acceder a la hoja de datos.")
            return None, None

        # Leer datos
        matriz = []
        for fila in ws.iter_rows(values_only=True):
            matriz.append(list(fila))

        numeros = matriz_a_numeros(matriz)
        logger.info(f"Archivo abierto exitosamente: {ruta} ({len(numeros)} números)")
        return numeros, ruta
        
    except Exception as e:
        logger.error(f"Error inesperado abriendo archivo: {type(e).__name__}: {e}")
        messagebox.showerror("Error", f"Error inesperado: {str(e)}")
        return None, None


def exportar_texto_plano(texto):
    """Exporta texto a archivo plano con validaciones."""
    if not isinstance(texto, str):
        logger.warning("Texto a exportar no válido: no es string")
        messagebox.showwarning("Aviso", "El texto a exportar no es válido.")
        return None
    
    if not texto.strip():
        logger.warning("Intento de exportar texto vacío")
        messagebox.showwarning("Aviso", "No hay texto decodificado para exportar.")
        return None

    nombre_base = simpledialog.askstring("Exportar", "Ingrese el nombre del archivo de texto:")
    if not nombre_base:
        logger.info("Exportación cancelada por usuario")
        return None

    valido, error_msg = validar_nombre_archivo(nombre_base)
    if not valido:
        logger.warning(f"Nombre rechazado en exportación: {error_msg}")
        messagebox.showerror("Error", error_msg)
        return None

    marca_tiempo = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_final = f"{Path(nombre_base).stem}_{marca_tiempo}.txt"

    ruta = filedialog.asksaveasfilename(
        title="Exportar texto",
        defaultextension=".txt",
        initialfile=nombre_final,
        filetypes=[("Archivos de texto", "*.txt")]
    )

    if not ruta:
        logger.info("Exportación cancelada")
        return None

    try:
        with open(ruta, "w", encoding="utf-8") as f:
            f.write(texto)
        logger.info(f"Texto exportado: {ruta} ({len(texto)} caracteres)")
        return ruta
    except IOError as e:
        logger.error(f"Error escribiendo archivo de texto: {e}")
        messagebox.showerror("Error", f"No se pudo exportar el archivo: {str(e)}")
        return None


# =========================================
# DIALOGO PERSONALIZADO
# =========================================

class NumeroDialog(tk.Toplevel):
    def __init__(self, parent, colores):
        super().__init__(parent)
        self.title("Decodificar")
        self.configure(bg=colores["panel"])
        self.resizable(False, False)
        self.resultado = None
        self.colores = colores

        self.transient(parent)
        self.grab_set()

        self.geometry("420x180")
        self.protocol("WM_DELETE_WINDOW", self.cancelar)

        cont = tk.Frame(self, bg=colores["panel"], padx=18, pady=18)
        cont.pack(fill="both", expand=True)

        titulo = tk.Label(
            cont,
            text="Decodificación",
            bg=colores["panel"],
            fg=colores["texto"],
            font=("Segoe UI", 12, "bold")
        )
        titulo.pack(anchor="w", pady=(0, 12))

        linea = tk.Frame(cont, bg=colores["panel"])
        linea.pack(anchor="w")

        tk.Label(
            linea,
            text="introduzca ",
            bg=colores["panel"],
            fg=colores["texto"],
            font=("Segoe UI", 11)
        ).pack(side="left")

        tk.Label(
            linea,
            text="el",
            bg=colores["panel"],
            fg=colores["fucsia"],
            font=("Segoe UI", 11, "bold")
        ).pack(side="left")

        tk.Label(
            linea,
            text=" número",
            bg=colores["panel"],
            fg=colores["texto"],
            font=("Segoe UI", 11)
        ).pack(side="left")

        self.entry = tk.Entry(
            cont,
            font=("Consolas", 12),
            bg=colores["panel2"],
            fg=colores["texto"],
            insertbackground=colores["fucsia"],
            relief="flat",
            width=28
        )
        self.entry.pack(fill="x", pady=(14, 16))
        self.entry.focus_set()

        botones = tk.Frame(cont, bg=colores["panel"])
        botones.pack(anchor="e")

        btn_ok = tk.Button(
            botones,
            text="Aceptar",
            command=self.aceptar,
            bg=colores["fucsia"],
            fg="white",
            relief="flat",
            padx=14,
            pady=6
        )
        btn_ok.pack(side="left", padx=6)

        btn_cancel = tk.Button(
            botones,
            text="Cancelar",
            command=self.cancelar,
            bg=colores["panel2"],
            fg=colores["texto"],
            relief="flat",
            padx=14,
            pady=6
        )
        btn_cancel.pack(side="left")

        self.bind("<Return>", lambda e: self.aceptar())
        self.bind("<Escape>", lambda e: self.cancelar())

        self.wait_window(self)

    def aceptar(self):
        self.resultado = self.entry.get().strip()
        self.destroy()

    def cancelar(self):
        self.resultado = None
        self.destroy()


# =========================================
# INTERFAZ
# =========================================

class App:
    def __init__(self, root):
        self.root = root
        self.root.title(f"{APP_NAME} - by ElSuinda @v@")
        self.root.geometry("1360x820")
        self.root.minsize(1150, 680)

        self.numeros_actuales = []
        self.ruta_excel_actual = None
        self.modificado = False

        self.colores = {
            "bg": "#07080b",
            "panel": "#10131a",
            "panel2": "#171b24",
            "panel3": "#1d2330",
            "texto": "#f4f7fb",
            "subtexto": "#aab3c2",
            "borde": "#262d3a",
            "fucsia": "#ff2da6",
            "fucsia2": "#d91b8f",
            "tabla_bg": "#0d1016",
            "tabla_head": "#1b2130",
            "verde": "#7ed957",
            "rojo": "#ff5e78"
        }

        self.configurar_icono()
        self.configurar_estilo()
        self.crear_interfaz()

    def configurar_icono(self):
        """Configura icono con validación de seguridad."""
        try:
            icono = os.path.join("assets", "icono.ico")
            if os.path.exists(icono) and os.path.isfile(icono):
                self.root.iconbitmap(icono)
                logger.info("Icono configurado exitosamente")
        except Exception as e:
            logger.warning(f"No se pudo cargar icono: {e}")

    def configurar_estilo(self):
        self.root.configure(bg=self.colores["bg"])
        style = ttk.Style()
        style.theme_use("clam")

        style.configure("TFrame", background=self.colores["bg"])
        style.configure("Card.TFrame", background=self.colores["panel"])

        style.configure(
            "Title.TLabel",
            background=self.colores["bg"],
            foreground=self.colores["texto"],
            font=("Segoe UI", 18, "bold")
        )

        style.configure(
            "Info.TLabel",
            background=self.colores["bg"],
            foreground=self.colores["subtexto"],
            font=("Segoe UI", 10)
        )

        style.configure(
            "CardTitle.TLabel",
            background=self.colores["panel"],
            foreground=self.colores["texto"],
            font=("Segoe UI", 12, "bold")
        )

        style.configure(
            "Dark.TButton",
            font=("Segoe UI", 10, "bold"),
            padding=10,
            background=self.colores["panel3"],
            foreground=self.colores["texto"],
            borderwidth=0
        )

        style.map(
            "Dark.TButton",
            background=[("active", self.colores["fucsia"])],
            foreground=[("active", "white")]
        )

        style.configure(
            "Treeview",
            background=self.colores["tabla_bg"],
            fieldbackground=self.colores["tabla_bg"],
            foreground=self.colores["texto"],
            rowheight=30,
            bordercolor=self.colores["borde"],
            borderwidth=1,
            font=("Consolas", 10)
        )

        style.configure(
            "Treeview.Heading",
            background=self.colores["tabla_head"],
            foreground=self.colores["texto"],
            font=("Segoe UI", 10, "bold"),
            relief="flat"
        )

    def crear_interfaz(self):
        # Configurar grid para la ventana raíz
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=0)  # Status bar
        self.root.rowconfigure(2, weight=0)  # Footer

        # Frame principal
        contenedor = ttk.Frame(self.root)
        contenedor.grid(row=0, column=0, sticky="nsew", padx=14, pady=14)
        contenedor.columnconfigure(0, weight=1)
        contenedor.rowconfigure(0, weight=0)  # Cabecera
        contenedor.rowconfigure(1, weight=0)  # Botones
        contenedor.rowconfigure(2, weight=1)  # Área de texto (expandible)

        # Cabecera
        cabecera = ttk.Frame(contenedor)
        cabecera.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        ttk.Label(cabecera, text=APP_NAME, style="Title.TLabel").pack(anchor="w")

        # Barra de botones
        barra = ttk.Frame(contenedor)
        barra.grid(row=1, column=0, sticky="ew", pady=(0, 12))

        ttk.Button(barra, text="Encriptar y Guardar", style="Dark.TButton", command=self.encriptar_y_guardar).pack(side="left", padx=5)
        ttk.Button(barra, text="Abrir y Desencriptar", style="Dark.TButton", command=self.abrir_y_desencriptar).pack(side="left", padx=5)
        ttk.Button(barra, text="Desencriptar", style="Dark.TButton", command=self.desencriptar_pegar).pack(side="left", padx=5)
        ttk.Button(barra, text="Copiar", style="Dark.TButton", command=self.copiar_contenido).pack(side="left", padx=5)
        ttk.Button(barra, text="Limpiar", style="Dark.TButton", command=self.limpiar).pack(side="left", padx=5)

        # Área de texto dividida (texto y números)
        texto_frame = ttk.Frame(contenedor, style="Card.TFrame", padding=10)
        texto_frame.grid(row=2, column=0, sticky="nsew")
        texto_frame.columnconfigure(0, weight=1)  # Texto
        texto_frame.columnconfigure(1, weight=1)  # Números
        texto_frame.rowconfigure(0, weight=0)     # Label
        texto_frame.rowconfigure(1, weight=1)     # Contenido

        # Labels para las dos secciones
        label_texto = ttk.Label(texto_frame, text="Texto Original", style="CardTitle.TLabel")
        label_texto.grid(row=0, column=0, sticky="ew", padx=(0, 5), pady=(0, 8))
        
        label_numeros = ttk.Label(texto_frame, text="Secuencia Numérica", style="CardTitle.TLabel")
        label_numeros.grid(row=0, column=1, sticky="ew", padx=(5, 0), pady=(0, 8))

        # Frame para texto
        texto_input_frame = ttk.Frame(texto_frame, style="Card.TFrame")
        texto_input_frame.grid(row=1, column=0, sticky="nsew", padx=(0, 5))
        texto_input_frame.columnconfigure(0, weight=1)
        texto_input_frame.rowconfigure(0, weight=1)

        self.texto_principal = tk.Text(
            texto_input_frame,
            wrap="word",
            font=("Consolas", 12),
            bg=self.colores["panel2"],
            fg=self.colores["texto"],
            insertbackground=self.colores["fucsia"],
            relief="flat",
            padx=12,
            pady=12
        )
        self.texto_principal.grid(row=0, column=0, sticky="nsew")
        self.texto_principal.bind("<KeyRelease>", self.actualizar_numeros)

        # Scrollbar para texto
        scroll_texto = ttk.Scrollbar(texto_input_frame, orient="vertical", command=self.texto_principal.yview)
        scroll_texto.grid(row=0, column=1, sticky="ns")
        self.texto_principal.config(yscrollcommand=scroll_texto.set)

        # Frame para números
        numeros_frame = ttk.Frame(texto_frame, style="Card.TFrame")
        numeros_frame.grid(row=1, column=1, sticky="nsew", padx=(5, 0))
        numeros_frame.columnconfigure(0, weight=1)
        numeros_frame.rowconfigure(0, weight=1)

        self.numeros_display = tk.Text(
            numeros_frame,
            wrap="word",
            font=("Consolas", 11),
            bg=self.colores["panel2"],
            fg=self.colores["verde"],
            relief="flat",
            padx=12,
            pady=12
        )
        self.numeros_display.grid(row=0, column=0, sticky="nsew")

        # Scrollbar para números
        scroll_numeros = ttk.Scrollbar(numeros_frame, orient="vertical", command=self.numeros_display.yview)
        scroll_numeros.grid(row=0, column=1, sticky="ns")
        self.numeros_display.config(yscrollcommand=scroll_numeros.set)

        # Status bar
        self.status_var = tk.StringVar(value="Listo.")
        status = tk.Label(
            self.root,
            textvariable=self.status_var,
            anchor="w",
            bg="#05060a",
            fg=self.colores["subtexto"],
            padx=10,
            pady=7
        )
        status.grid(row=1, column=0, sticky="ew")

        # Footer con crédito
        footer = tk.Label(
            self.root,
            text="by ElSuinda @v@",
            anchor="e",
            bg="#05060a",
            fg=self.colores["borde"],
            padx=10,
            pady=5,
            font=("Consolas", 8)
        )
        footer.grid(row=2, column=0, sticky="ew")

    def set_status(self, texto):
        self.status_var.set(texto)

    def actualizar_numeros(self, event=None):
        """Actualiza la visualización de números cuando cambia el texto."""
        try:
            texto = self.texto_principal.get("1.0", "end-1c")
            
            # Generar números
            self.numeros_actuales = texto_a_numeros(texto)
            
            # Mostrar números en el widget
            self.numeros_display.delete("1.0", "end")
            
            if self.numeros_actuales:
                numeros_str = " ".join(str(n) for n in self.numeros_actuales)
                self.numeros_display.insert("1.0", numeros_str)
            
        except Exception as e:
            logger.error(f"Error actualizando números: {e}")

    def encriptar_y_guardar(self):
        """Encripta el texto y lo guarda en Excel."""
        try:
            texto = self.texto_principal.get("1.0", "end-1c")
            
            if not texto.strip():
                messagebox.showwarning("Aviso", "Por favor ingrese un texto para encriptar.")
                return
            
            # Encriptar
            numeros = texto_a_numeros(texto)
            self.numeros_actuales = numeros
            
            # Guardar
            nombre_base = simpledialog.askstring("Guardar", "Ingrese el nombre del archivo:")
            if not nombre_base:
                return
            
            valido, error_msg = validar_nombre_archivo(nombre_base)
            if not valido:
                messagebox.showerror("Error", error_msg)
                return
            
            nombre_final = construir_nombre_archivo(nombre_base)
            ruta = filedialog.asksaveasfilename(
                title="Guardar archivo",
                defaultextension=".xlsx",
                initialfile=nombre_final,
                filetypes=[("Archivos Excel", "*.xlsx")]
            )
            
            if not ruta:
                return
            
            escribir_excel(ruta, numeros)
            self.ruta_excel_actual = ruta
            
            self.texto_principal.delete("1.0", "end")
            self.set_status(f"Archivo guardado: {ruta}")
            messagebox.showinfo("Guardado", "Mensaje encriptado y guardado correctamente.")
            logger.info(f"Mensaje encriptado y guardado: {ruta}")
            
        except Exception as e:
            logger.error(f"Error encriptando y guardando: {e}")
            messagebox.showerror("Error", f"Error: {str(e)}")

    def abrir_y_desencriptar(self):
        """Abre un archivo Excel y desencripta con la clave."""
        try:
            datos = abrir_desde_excel()
            if not datos or datos == (None, None):
                return
            
            numeros, ruta = datos
            if not isinstance(numeros, list):
                messagebox.showerror("Error", "Error al leer el archivo.")
                return
            
            self.numeros_actuales = numeros
            self.ruta_excel_actual = ruta
            
            # Pedir clave
            clave_dialog = simpledialog.askstring("Clave", "Ingrese la clave (1.618):", show="*")
            if not clave_dialog:
                return
            
            # Desencriptar
            ok, texto = numeros_a_texto(numeros, clave_dialog)
            
            self.texto_principal.delete("1.0", "end")
            self.texto_principal.insert("1.0", texto)
            
            if ok:
                logger.info("Desencriptación exitosa")
                self.set_status("Desencriptación exitosa.")
                messagebox.showinfo("Exito", "Mensaje desencriptado correctamente.")
            else:
                logger.warning("Desencriptación fallida: clave incorrecta")
                self.set_status("Error: clave incorrecta.")
                messagebox.showwarning("Error", texto)
                
        except Exception as e:
            logger.error(f"Error abriendo archivo: {e}")
            messagebox.showerror("Error", f"Error: {str(e)}")

    def desencriptar_pegar(self):
        """Desencripta números pegados o del panel de secuencia."""
        try:
            # Intentar obtener números del panel de secuencia primero
            numeros_texto = self.numeros_display.get("1.0", "end-1c").strip()
            
            # Si el panel está vacío, obtener del portapapeles
            if not numeros_texto:
                self.root.update()
                try:
                    numeros_texto = self.root.clipboard_get()
                except Exception:
                    messagebox.showwarning("Aviso", "No hay contenido en el portapapeles ni en el panel.")
                    return
            
            if not numeros_texto.strip():
                messagebox.showwarning("Aviso", "El portapapeles está vacío.")
                return
            
            # Parsear números
            try:
                numeros = [int(n.strip()) for n in numeros_texto.replace('\n', ' ').split() if n.strip()]
            except ValueError:
                messagebox.showerror("Error", "Los números no son válidos. Use espacios o saltos de línea.")
                return
            
            if not numeros:
                messagebox.showwarning("Aviso", "No se encontraron números válidos.")
                return
            
            self.numeros_actuales = numeros
            
            # Pedir clave
            clave_dialog = simpledialog.askstring("Clave", "Ingrese la clave (1.618):", show="*")
            if not clave_dialog:
                return
            
            # Desencriptar
            ok, texto = numeros_a_texto(numeros, clave_dialog)
            
            self.texto_principal.delete("1.0", "end")
            self.texto_principal.insert("1.0", texto)
            
            if ok:
                logger.info("Desencriptación exitosa desde portapapeles")
                self.set_status("Desencriptación exitosa.")
                messagebox.showinfo("Exito", "El número es áureo φ")
            else:
                logger.warning("Desencriptación fallida: clave incorrecta")
                self.set_status("Error: clave incorrecta.")
                messagebox.showwarning("Error", texto)
                
        except Exception as e:
            logger.error(f"Error en desdeciframiento: {e}")
            messagebox.showerror("Error", f"Error: {str(e)}")

    def copiar_contenido(self):
        """Copia los números del panel de secuencia al portapapeles."""
        try:
            texto = self.numeros_display.get("1.0", "end-1c")
            if not texto.strip():
                messagebox.showwarning("Aviso", "No hay números para copiar.")
                return
            
            self.root.clipboard_clear()
            self.root.clipboard_append(texto)
            self.root.update()
            
            logger.info(f"Contenido copiado al portapapeles ({len(texto)} caracteres)")
            self.set_status("Contenido copiado al portapapeles.")
            messagebox.showinfo("Copiado", "Contenido copiado correctamente.")
            
        except Exception as e:
            logger.error(f"Error copiando contenido: {e}")
            messagebox.showerror("Error", f"Error: {str(e)}")

    def limpiar(self):
        """Limpia el área de texto y la secuencia numérica."""
        try:
            self.texto_principal.delete("1.0", "end")
            self.numeros_display.delete("1.0", "end")
            self.numeros_actuales = []
            self.ruta_excel_actual = None
            self.set_status("Interfaz limpiada.")
            logger.info("Interfaz limpiada")
        except Exception as e:
            logger.error(f"Error limpiando: {e}")
            self.set_status("Error limpiando la interfaz.")




if __name__ == "__main__":
    try:
        logger.info("====== INICIANDO APLICACION SEPTEM ======")
        root = tk.Tk()
        app = App(root)
        logger.info("Aplicación iniciada exitosamente")
        root.mainloop()
    except Exception as e:
        logger.critical(f"Error crítico iniciando aplicación: {e}", exc_info=True)
        print(f"Error crítico: {e}")
    finally:
        logger.info("====== CERRANDO APLICACION SEPTEM ======")
